"""
Scrape l'historique de commandes Rexel pro et enrichit le fichier Excel
scripts/data/Bibliothèque matériaux application Profero.xlsx.

Workflow :
1. Ouvre la page Mes commandes avec le cookie de session.
2. Élargit le filtre de date à 12 mois et pageSize à 50.
3. Parcourt toutes les pages, collecte les URLs des commandes.
4. Visite chaque commande, extrait les lignes (code, nom, ref Rexel, prix, image).
5. Déduplique par code produit, exclut les refs déjà présentes dans l'Excel.
6. Sauvegarde l'Excel (backup auto), append les nouvelles lignes.

Usage :
    python scripts/import_rexel.py                    # tout, browser invisible
    python scripts/import_rexel.py --visible          # browser visible
    python scripts/import_rexel.py --dry-run          # n'écrit pas l'Excel
    python scripts/import_rexel.py --max-orders 3     # limite (test rapide)
    python scripts/import_rexel.py --months 6         # filtre N mois (def 12)

Configuration (.env à la racine) :
    REXEL_COOKIES_HEADER=<header Cookie complet copié depuis DevTools>
"""

import argparse
import re
import shutil
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import unquote

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "scripts" / "data"
DATA_FILE = DATA_DIR / "Bibliothèque matériaux application Profero.xlsx"
SHEET_NAME = "bibliotheque_materiaux"
FOURNISSEUR = "Rexel"
BASE_URL = "https://www.rexel.fr"
ORDERS_URL = f"{BASE_URL}/frx/my-account/orders"

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)


# ─── .env ─────────────────────────────────────────────────────────────────────
def load_env():
    env_path = ROOT / ".env"
    env = {}
    if not env_path.exists():
        return env
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, val = line.split("=", 1)
        env[key.strip()] = val.strip().strip('"').strip("'")
    return env


def build_cookies(env):
    raw = env.get("REXEL_COOKIES_HEADER")
    if raw:
        cookies = []
        for pair in raw.split(";"):
            pair = pair.strip()
            if "=" not in pair:
                continue
            name, value = pair.split("=", 1)
            cookies.append({
                "name": name.strip(),
                "value": value.strip(),
                "domain": ".rexel.fr",
                "path": "/",
            })
        if cookies:
            return cookies
    fallback = env.get("REXEL_SESSION_COOKIE")
    if fallback:
        return [{
            "name": "JSESSIONID",
            "value": fallback,
            "domain": ".rexel.fr",
            "path": "/",
        }]
    sys.exit("ERREUR : aucun cookie dans .env (REXEL_COOKIES_HEADER attendu)")


# ─── Catégorie extraite depuis l'URL produit ──────────────────────────────────
# URL type : /frx/Catégorie/<niv1>/<niv2>/<niv3>/<nom-slug>/<refRexel>/p/<code>
# On garde le niveau 1 comme catégorie (compatible avec les valeurs existantes
# du fichier : "Electricité", "Plomberie"…).
def extract_categorie_from_url(url):
    if not url:
        return ""
    decoded = unquote(url)
    m = re.search(r"/Cat[eé]gorie/([^/]+)/", decoded)
    return m.group(1).replace("-", " ").strip() if m else ""


def extract_ref_rexel_from_url(url):
    if not url:
        return None
    m = re.search(r"/([^/]+)/p/\d+/?$", url)
    return m.group(1) if m else None


def parse_price(s):
    if s is None:
        return None
    # "38,28485 €" / "1 234,56 €" / "38.28"
    cleaned = re.sub(r"[^\d,.\-]", "", str(s)).replace(",", ".")
    if cleaned.count(".") > 1:
        # Cas "1.234.56" : on garde le dernier point comme décimale
        parts = cleaned.split(".")
        cleaned = "".join(parts[:-1]) + "." + parts[-1]
    try:
        return float(cleaned)
    except ValueError:
        return None


# ─── Excel ────────────────────────────────────────────────────────────────────
def load_existing_refs():
    if not DATA_FILE.exists():
        return set()
    wb = load_workbook(DATA_FILE, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        return set()
    ws = wb[SHEET_NAME]
    refs = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is not None:
            refs.add(str(row[1]).strip())
    return refs


def backup_excel():
    if not DATA_FILE.exists():
        return None
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = DATA_FILE.with_name(f"{DATA_FILE.stem}.backup-{ts}.xlsx")
    shutil.copy2(DATA_FILE, backup_path)
    return backup_path


def append_articles(articles):
    """Ajoute les articles à la fin de la feuille bibliotheque_materiaux."""
    wb = load_workbook(DATA_FILE)
    ws = wb[SHEET_NAME]
    # Colonnes : nom, reference, fournisseur, categorie, prix_unitaire, unite,
    #            stock_min, lien_fournisseur, photo_url, notes
    for a in articles:
        ws.append([
            a["nom"],
            a["reference"],
            FOURNISSEUR,
            a.get("categorie", ""),
            a.get("prix_unitaire"),
            "U",
            0,
            a.get("lien_fournisseur"),
            a.get("photo_url"),
            a.get("notes"),
        ])
    wb.save(DATA_FILE)


# ─── Scraping ─────────────────────────────────────────────────────────────────
def set_date_filter(page, months_back):
    """Élargit le filtre de date à N mois en arrière, change pageSize à 50,
    déclenche le rechargement de la liste."""
    end_date = datetime.now()
    start_date = end_date - timedelta(days=months_back * 31)
    start_str = start_date.strftime("%d.%m.%Y")
    end_str = end_date.strftime("%d.%m.%Y")
    print(f"  → filtre date : {start_str} → {end_str}, pageSize=50")
    # Le bouton "Rechercher" déclenche un AJAX vers getOrderHistory.ajax.
    # On modifie les valeurs des inputs cachés puis on appelle directement
    # la fonction JS exposée par Rexel si possible, sinon on submit le form.
    page.evaluate(
        """
        (params) => {
            const sd = document.getElementById('orderStartDate');
            const ed = document.getElementById('orderEndDate');
            if (sd) sd.value = params.start;
            if (ed) ed.value = params.end;
            // Change pageSize si un select existe
            const sel = document.querySelector('select[name="pageSize"], select#pageSize, select.pageSize');
            if (sel) {
                sel.value = '50';
                sel.dispatchEvent(new Event('change', { bubbles: true }));
            }
        }
        """,
        {"start": start_str, "end": end_str},
    )
    # Clic sur le bouton de recherche / appliquer
    # On essaie plusieurs sélecteurs (Rexel a plusieurs boutons possibles)
    clicked = False
    for sel in [
        "button#searchOrderBtn",
        "button.searchOrderBtn",
        "a.searchOrderBtn",
        "input.searchOrderBtn",
        "button.orderHistorySearchBtn",
        "#applyFilter",
    ]:
        try:
            btn = page.query_selector(sel)
            if btn:
                btn.click()
                clicked = True
                break
        except Exception:
            pass
    if not clicked:
        # Fallback : recharger la page avec URL params (Hybris souvent ok)
        url = (
            f"{ORDERS_URL}?"
            f"orderStartDate={start_str}&orderEndDate={end_str}&pageSize=50"
        )
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass


def get_total_pages(page):
    """Lit le nombre total de pages depuis la pagination Bootpag."""
    return page.evaluate(
        """
        () => {
            const lis = document.querySelectorAll('.OrderHistoryPgnTion li[data-lp]');
            let max = 1;
            lis.forEach(li => {
                const n = parseInt(li.dataset.lp);
                if (!isNaN(n) && n > max) max = n;
            });
            return max;
        }
        """
    )


def collect_orders_on_page(page):
    """Renvoie les URLs des commandes affichées sur la page courante."""
    return page.evaluate(
        f"""
        () => Array.from(document.querySelectorAll('div.orderHistory-container'))
            .map(div => {{
                const id = div.id;
                const hidden = div.querySelector('input.myAccountOrderDetailsUrl');
                const url = hidden ? hidden.value.replace(/\\?$/, '') : null;
                return {{ id, url }};
            }})
            .filter(o => o.url)
        """
    )


def go_to_page(page, page_num):
    """Clique sur la page N de la pagination Bootpag, attend rechargement."""
    page.evaluate(
        f"""
        () => {{
            const li = document.querySelector('.OrderHistoryPgnTion li[data-lp="{page_num}"]');
            const a = li ? li.querySelector('a') : null;
            if (a) a.click();
        }}
        """
    )
    # Bootpag fait un AJAX, on attend un peu et networkidle
    time.sleep(1.5)
    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass


def extract_lines_from_order(page, order_url):
    """Visite la page détail d'une commande, extrait les lignes produit."""
    full_url = BASE_URL + order_url if order_url.startswith("/") else order_url
    page.goto(full_url, wait_until="domcontentloaded", timeout=30000)
    try:
        page.wait_for_load_state("networkidle", timeout=10000)
    except Exception:
        pass
    return page.evaluate(
        r"""
        () => {
            const rows = Array.from(document.querySelectorAll('tr.orderItem'));
            return rows.map(tr => {
                const code = tr.dataset.productCode || '';
                const cb = tr.querySelector('input.subOrderSelectCheckbox');
                const name = (cb?.dataset.productname || '').replace(/\s+/g, ' ').trim();
                const brand = (cb?.dataset.productbrand || '').trim();
                const priceStr = cb?.dataset.productprice || '';
                // Le lien produit principal
                const link = tr.querySelector('.item-Name a, td.order-img a[href*="/p/"]');
                const url = link ? link.getAttribute('href') : null;
                return { code, name, brand, priceStr, url };
            }).filter(r => r.code);
        }
        """
    )


def scrape_orders(args):
    env = load_env()
    cookies = build_cookies(env)
    existing_refs = load_existing_refs()
    print(f"[biblio] {len(existing_refs)} refs déjà présentes")

    all_articles_by_code = {}

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=not args.visible)
        context = browser.new_context(
            user_agent=UA, viewport={"width": 1440, "height": 900}, locale="fr-FR",
        )
        context.add_cookies(cookies)
        page = context.new_page()

        print(f"\n[phase 1] accès à {ORDERS_URL}")
        page.goto(ORDERS_URL, wait_until="domcontentloaded", timeout=30000)
        try:
            page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            pass

        if "/login" in page.url.lower():
            sys.exit("[ERREUR] Cookie invalide ou expiré — recopier depuis DevTools.")

        print(f"[phase 1] connecté, url = {page.url}")

        print(f"\n[phase 2] application du filtre {args.months} mois")
        set_date_filter(page, args.months)

        total_pages = get_total_pages(page)
        print(f"[phase 2] {total_pages} page(s) de commandes")

        # Collecte URLs commandes
        orders = []
        for p_num in range(1, total_pages + 1):
            if p_num > 1:
                print(f"  → page {p_num}")
                go_to_page(page, p_num)
            page_orders = collect_orders_on_page(page)
            orders.extend(page_orders)

        print(f"\n[phase 3] {len(orders)} commandes trouvées")
        if args.max_orders:
            orders = orders[: args.max_orders]
            print(f"           limité à {len(orders)} via --max-orders")

        # Parcours détails
        for i, order in enumerate(orders, 1):
            print(f"  [{i:>3}/{len(orders)}] commande {order['id']}", end=" -> ")
            try:
                lines = extract_lines_from_order(page, order["url"])
            except Exception as e:
                print(f"[erreur] {e}")
                continue
            new_in_order = 0
            for line in lines:
                code = line["code"]
                if not code or code in all_articles_by_code:
                    continue
                ref_rexel = extract_ref_rexel_from_url(line["url"])
                ref = ref_rexel or code  # fallback : code interne si pas de réf
                if ref in existing_refs:
                    continue
                prix = parse_price(line["priceStr"])
                cat = extract_categorie_from_url(line["url"])
                lien = (BASE_URL + line["url"]) if line["url"] and line["url"].startswith("/") else line["url"]
                all_articles_by_code[code] = {
                    "nom": line["name"],
                    "reference": ref,
                    "categorie": cat,
                    "prix_unitaire": prix,
                    "lien_fournisseur": lien,
                    "photo_url": f"https://api.rexel.fr/media/api/fr/content/{code}/LARGE",
                    "notes": f"Marque: {line['brand']}" if line["brand"] else None,
                }
                new_in_order += 1
            print(f"{len(lines)} lignes ({new_in_order} nouvelles)")
            # Throttle pour éviter d'être flag bot
            time.sleep(0.8)

        browser.close()

    articles = list(all_articles_by_code.values())
    print(f"\n[total] {len(articles)} articles uniques à ajouter (hors doublons / refs existantes)")
    if args.dry_run:
        print("\n[dry-run] pas d'écriture. Aperçu des 5 premiers :")
        for a in articles[:5]:
            prix = a["prix_unitaire"] if a["prix_unitaire"] is not None else 0.0
            print(f"  - {a['reference']:>15} | {a['nom'][:60]:60} | {prix:>8.2f} EUR")
        return

    if not articles:
        print("Rien à ajouter.")
        return

    backup_path = backup_excel()
    if backup_path:
        print(f"\n[backup] {backup_path.name}")
    append_articles(articles)
    print(f"[ok] {len(articles)} lignes ajoutées à {DATA_FILE.name}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scraper Rexel pro → biblio Excel")
    parser.add_argument("--visible", action="store_true",
                        help="Browser visible (défaut : invisible)")
    parser.add_argument("--dry-run", action="store_true",
                        help="N'écrit pas dans l'Excel, affiche juste un aperçu")
    parser.add_argument("--max-orders", type=int, default=None,
                        help="Limite le nb de commandes parcourues (test rapide)")
    parser.add_argument("--months", type=int, default=12,
                        help="Fenêtre de date en mois (défaut 12)")
    args = parser.parse_args()
    scrape_orders(args)
